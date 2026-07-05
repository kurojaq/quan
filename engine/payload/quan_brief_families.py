#!/usr/bin/env python3
"""
quan_brief_families.py — renders the §3.7 full graph families from a recalced oracle
workbook (golden_reference recalced ATM-centered with the session chain, i.e. the
_tensor_recalc.xlsx that quan_tensor.inject_pressure+recalc already produces).

Reproduces the partner's complete chart set:
  * SOP wave field (headline)         -> _compass_wavefield.png
  * SOP panel set (6 panels)          -> _sop_panels.png
  * TSC Curvature family (21 charts)  -> _tsc_curvature.png
  * TSC Gradient family (23 charts)   -> _tsc_gradient.png
  * Hessian curvature-geometry block  -> returned dict (rendered as §3.7 text)

All families are oracle-sourced VISUALIZATION, not new signals.
"""
import os, re, json
import numpy as np
import openpyxl
from openpyxl.utils import column_index_from_string
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt

BG = "#0d0f14"; FG = "#c9cdd6"; GRID = "#2a2f3a"; CYAN = "#3fb6d3"
CUR_SHEET = "Time State Compass(Curvature)"
GRAD_SHEET = "Time State Compass(Gradient)"
SOP_SHEET = "SOP Folding"


def _smooth(xs, ys, n=300):
    """Spline-interpolate (Excel smoothed-line look) so the headline curve reads
    wavy like the §3.5 compass graphs. Linear fallback if scipy is unavailable.
    Draws the SAME computed points with a smooth curve between samples — adds no data."""
    x = np.asarray(xs, float); y = np.asarray(ys, float)
    m = np.isfinite(x) & np.isfinite(y)
    x, y = x[m], y[m]
    if len(x) < 2:
        return x, y
    order = np.argsort(x); x, y = x[order], y[order]
    ux, idx = np.unique(x, return_index=True); x, y = ux, y[idx]
    if len(x) < 4:
        return x, y
    xs2 = np.linspace(x.min(), x.max(), n)
    try:
        from scipy.interpolate import make_interp_spline
        return xs2, make_interp_spline(x, y, k=3)(xs2)
    except Exception:
        return xs2, np.interp(xs2, x, y)


def _parse_ref(ref):
    m = re.match(r"'?(.*?)'?!\$?([A-Z]+)\$?(\d+)(?::\$?([A-Z]+)\$?(\d+))?", ref or "")
    if not m:
        return None
    sheet = m.group(1)
    c1 = column_index_from_string(m.group(2)); r1 = int(m.group(3))
    r2 = int(m.group(5)) if m.group(5) else r1
    return sheet, c1, r1, r2


def _series_xy(ws, xref, yref, stoprow):
    px = _parse_ref(xref); py = _parse_ref(yref)
    if not py:
        return None, None
    col = py[1]; r1 = py[2]
    xs, ys = [], []
    xcol = px[1] if px else None
    for r in range(r1, stoprow + 1):
        yv = ws.cell(r, col).value
        if not isinstance(yv, (int, float)):
            continue
        xv = ws.cell(r, xcol).value if xcol else (r - r1)
        if not isinstance(xv, (int, float)):
            continue
        xs.append(float(xv)); ys.append(float(yv))
    return xs, ys


def _data_stoprow(ws, paircol=3, hardcap=23):
    """last row where the Pair/chronometer column is numeric (before the Hessian block)."""
    last = 1
    for r in range(2, hardcap + 1):
        v = ws.cell(r, paircol).value
        if isinstance(v, (int, float)):
            last = r
        else:
            break
    return last


def _style(ax):
    ax.set_facecolor(BG)
    for s in ax.spines.values():
        s.set_color(GRID)
    ax.tick_params(colors=FG, labelsize=5.5, length=2)
    ax.axhline(0, color="#5a6172", lw=0.6, zorder=1)
    ax.grid(True, color=GRID, lw=0.35, alpha=0.5)


_SKIP_HDR = {"pair", "chronometer watch", "chronometer", ""}


def _find_paircol(ws):
    for c in range(1, ws.max_column + 1):
        h = ws.cell(1, c).value
        if isinstance(h, str) and h.strip().lower() == "pair":
            return c
    return 3  # default col C


def _render_family(wb, allowed_titles, sheet_name, out_png, suptitle):
    """Header-driven, constrained to the canonical chart titles for this family."""
    if sheet_name not in wb.sheetnames:
        return None
    ws = wb[sheet_name]
    paircol = _find_paircol(ws)
    stop = _data_stoprow(ws, paircol=paircol)
    xs = [ws.cell(r, paircol).value for r in range(2, stop + 1)]

    def _norm(s):
        return re.sub(r"[^a-z0-9]", "", str(s).lower())
    allow = {_norm(t): t for t in allowed_titles}
    # collect headed numeric columns
    found = {}
    for c in range(1, ws.max_column + 1):
        if c == paircol:
            continue
        h = ws.cell(1, c).value
        if not isinstance(h, str) or h.strip().lower() in _SKIP_HDR:
            continue
        key = _norm(h)
        if key not in allow or key in found:
            continue
        ys = [ws.cell(r, c).value for r in range(2, stop + 1)]
        nums = [(x, y) for x, y in zip(xs, ys) if isinstance(x, (int, float)) and isinstance(y, (int, float))]
        if len(nums) >= 2:
            found[key] = (h.strip(), nums)
    # order by the canonical title order
    panels = [found[_norm(t)] for t in allowed_titles if _norm(t) in found]
    if not panels:
        return None
    n = len(panels); cols = 4; rows = (n + cols - 1) // cols
    fig, axes = plt.subplots(rows, cols, figsize=(11, 2.0 * rows), facecolor=BG)
    axes = axes.flatten() if hasattr(axes, "flatten") else [axes]
    for i, (title, nums) in enumerate(panels):
        ax = axes[i]; _style(ax)
        xx = [p[0] for p in nums]; yy = [p[1] for p in nums]
        ax.plot(xx, yy, color=CYAN, lw=1.3, marker="o", ms=2.0, mfc=CYAN, mec=CYAN)
        ax.set_title(title[:34], color=FG, fontsize=6.2, pad=3)
    for j in range(n, len(axes)):
        axes[j].set_visible(False)
    fig.suptitle(f"{suptitle}  ({n} charts)", color=FG, fontsize=9, y=0.997)
    fig.tight_layout(rect=[0, 0, 1, 0.985])
    fig.savefig(out_png, dpi=120, facecolor=BG); plt.close(fig)
    return out_png


def _sop_col(ws, header, stop, paircol=5):
    """find a SOP column by its row-1 header; return (pair_x, y) numeric pairs."""
    target = None
    for c in range(1, ws.max_column + 1):
        h = ws.cell(1, c).value
        if isinstance(h, str) and h.strip().lower() == header.strip().lower():
            target = c; break
    if target is None:
        return None, None
    xs, ys = [], []
    for r in range(2, stop + 1):
        yv = ws.cell(r, target).value; xv = ws.cell(r, paircol).value
        if isinstance(yv, (int, float)) and isinstance(xv, (int, float)):
            xs.append(float(xv)); ys.append(float(yv))
    return xs, ys


def _render_sop(wb, out_wavefield, out_panels):
    if SOP_SHEET not in wb.sheetnames:
        return []
    ws = wb[SOP_SHEET]
    stop = _data_stoprow(ws, paircol=5)
    out = []
    # headline wave field: product (white) + curvature (red) + gradient (gold) + envelope (green)
    series = [("SOPG*SOPC", "#ffffff", "product"),
              ("SOP Curvature", "#ef5350", "curvature"),
              ("SOP Gradient", "#ffb300", "gradient"),
              ("Product tension", "#26a69a", "tension/envelope")]
    fig, ax = plt.subplots(figsize=(12.5, 5.6), dpi=130, facecolor=BG); _style(ax)
    any_plotted = False
    for hdr, col, lab in series:
        xs, ys = _sop_col(ws, hdr, stop)
        if xs and len(xs) > 1:
            sx, sy = _smooth(xs, ys)        # spline-smoothed, like the §3.5 compass graphs
            ax.plot(sx, sy, color=col, lw=2.4 if lab == "product" else 1.5, label=lab)
            any_plotted = True
    ax.axhline(0, color=GRID, lw=0.9)       # zero reference so fold depth reads clearly
    ax.set_title("SOP wave field (headline) — product = fold/coherence wave", color=FG, fontsize=12)
    ax.set_xlabel("chronometer watch  (0 ATM \u2192 1 expiry arc)", color=FG, fontsize=9)
    ax.tick_params(colors=FG, labelsize=8)
    if any_plotted:
        lg = ax.legend(fontsize=8, facecolor=BG, edgecolor=GRID, labelcolor=FG, loc="lower left")
    fig.tight_layout(); fig.savefig(out_wavefield, dpi=130, facecolor=BG); plt.close(fig)
    out.append(out_wavefield)
    # panel set: product, product tension, product curvature, SOPG/SOPC, SOPC/SOPG, raw SOPG & SOPC
    panels = [("SOPG*SOPC", "product (fold/coherence)"),
              ("Product tension", "Product Tension (J+next)"),
              ("Product Curvature", "Product Curvature"),
              ("SOPG/SOPC", "SOPG / SOPC (ratio)"),
              ("SOPC/SOPG", "SOPC / SOPG (inverse)"),
              (("SOP Gradient", "SOP Curvature"), "SOPG & SOPC (raw factors)")]
    fig, axes = plt.subplots(2, 3, figsize=(11, 5.2), facecolor=BG)
    axes = axes.flatten()
    for i, (hdr, title) in enumerate(panels):
        ax = axes[i]; _style(ax)
        if isinstance(hdr, tuple):
            cols = ["#ffb300", "#ef5350"]
            for h, c in zip(hdr, cols):
                xs, ys = _sop_col(ws, h, stop)
                if xs and len(xs) > 1:
                    ax.plot(xs, ys, color=c, lw=1.2, label=h, marker="o", ms=2)
            ax.legend(fontsize=5.5, facecolor=BG, edgecolor=GRID, labelcolor=FG)
        else:
            xs, ys = _sop_col(ws, hdr, stop)
            if xs and len(xs) > 1:
                ax.plot(xs, ys, color="#ffffff" if i == 0 else CYAN, lw=1.4, marker="o", ms=2)
            else:
                ax.text(0.5, 0.5, "(flat / sparse)", color="#5a6172", fontsize=6, ha="center", va="center", transform=ax.transAxes)
        ax.set_title(title, color=FG, fontsize=7, pad=3)
    fig.suptitle("SOP Folding — full panel set  (shared chronometer x-axis)", color=FG, fontsize=9, y=0.995)
    fig.tight_layout(rect=[0, 0, 1, 0.97]); fig.savefig(out_panels, dpi=120, facecolor=BG); plt.close(fig)
    out.append(out_panels)
    return out


def read_hessian(wb):
    if CUR_SHEET not in wb.sheetnames:
        return {}
    ws = wb[CUR_SHEET]
    out = {}
    for r in range(20, 40):
        a = ws.cell(r, 1).value; b = ws.cell(r, 2).value
        if isinstance(a, str) and b is not None and not isinstance(b, str):
            out[a.strip()] = b
        elif isinstance(a, str) and isinstance(b, str) and "character" in a.lower():
            out[a.strip()] = b
    # also grab the character string
    for r in range(20, 40):
        a = ws.cell(r, 1).value; b = ws.cell(r, 2).value
        if isinstance(a, str) and "character" in a.lower():
            out["Curvature character"] = b
    return out


def render_families(recalc_xlsx, specs_json, out_prefix):
    """Returns (list_of_png_paths, hessian_dict). out_prefix like '/dir/brief_graph'."""
    wb = openpyxl.load_workbook(recalc_xlsx, data_only=True)
    specs = json.load(open(specs_json))
    titles = [s.get("title", "") for s in specs]
    # canonical groupings (chart order): 0-1 Dual Phase, 2-22 Curvature(21), 23-45 Gradient(23), 46-58 SOP(13)
    cur_titles = titles[2:23]
    grad_titles = titles[23:46]
    paths = []
    sop = _render_sop(wb, out_prefix + "_compass_wavefield.png", out_prefix + "_sop_panels.png")
    paths += sop
    cur = _render_family(wb, cur_titles, CUR_SHEET, out_prefix + "_tsc_curvature.png",
                         "Time State Compass — Curvature family (shared chronometer axis)")
    if cur: paths.append(cur)
    grad = _render_family(wb, grad_titles, GRAD_SHEET, out_prefix + "_tsc_gradient.png",
                          "Time State Compass — Gradient family (shared chronometer axis)")
    if grad: paths.append(grad)
    hess = read_hessian(wb)
    return paths, hess


if __name__ == "__main__":
    import sys
    p, h = render_families(sys.argv[1], sys.argv[2], sys.argv[3] if len(sys.argv) > 3 else "/tmp/fam")
    print("rendered:", p)
    print("hessian:", h)
