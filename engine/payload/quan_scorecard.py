"""
quan_scorecard.py — canonical Deep Strike per-strike block + 5-criterion scorecard.

Per-strike block is a cell-for-cell port of golden_reference Book_Strike_Level:
  K = PutLast-CallLast      L = PutVol-CallVol       M = PutOI/CallOI
  N = PutVol/CallVol        O = PutOI-CallOI          AN = PutPrem-CallPrem
  P = O/M                   Q = L/N                   R = O/L      S = L/O   T = P/Q (LR)
  W = KURT(L..T)            X = SKEW(L..T)            Mass = W/X   Force = X/W
  AP = AN/O (DID)           AR = AN/L (DIT)           AT = AP/AR (DR3)
  Y = 1+W_this+W_next (ICF) Z = Y/86400               STD_grad = Z_next-Z_this
Excel KURT/SKEW are sample (blank/error cells excluded), matching IFERROR(...,"").
Mass is canonical Kurt/Skew — NOT the alpha-beta-gamma proxy.
"""
import numpy as np, pandas as pd

# ---- Excel sample skew/kurt (blank-excluding) ----
def _clean_vec(vals):
    return [v for v in vals if isinstance(v,(int,float)) and np.isfinite(v)]

def excel_skew(vals):
    x=_clean_vec(vals); n=len(x)
    if n<3: return np.nan
    m=sum(x)/n; s=(sum((xi-m)**2 for xi in x)/(n-1))**0.5
    if s==0: return np.nan
    return n/((n-1)*(n-2))*sum(((xi-m)/s)**3 for xi in x)

def excel_kurt(vals):
    x=_clean_vec(vals); n=len(x)
    if n<4: return np.nan
    m=sum(x)/n; s=(sum((xi-m)**2 for xi in x)/(n-1))**0.5
    if s==0: return np.nan
    a=n*(n+1)/((n-1)*(n-2)*(n-3)); b=3*(n-1)**2/((n-2)*(n-3))
    return a*sum(((xi-m)/s)**4 for xi in x)-b

def _div(a,b):
    return a/b if (b not in (0,None) and np.isfinite(b) and np.isfinite(a)) else np.nan

def per_strike_block(frame):
    """frame: strike, callPrem, putPrem, callOI, putOI, callVol, putVol, callLatest, putLatest
    Returns a DataFrame with the canonical per-strike columns (ascending strike order)."""
    f=frame.copy()
    f["strike"]=pd.to_numeric(f["strike"],errors="coerce")
    f=f.dropna(subset=["strike"]).sort_values("strike").reset_index(drop=True)
    g=lambda c: pd.to_numeric(f.get(c), errors="coerce").fillna(0.0).values
    cP,pP=g("callPrem"),g("putPrem"); cOI,pOI=g("callOI"),g("putOI")
    cV,pV=g("callVol"),g("putVol");   cL,pL=g("callLatest"),g("putLatest")
    n=len(f); rows=[]
    for i in range(n):
        K=_div(pL[i]-cL[i],1) if True else np.nan          # I-H (never errors here)
        K=pL[i]-cL[i]
        L=pV[i]-cV[i]
        M=_div(pOI[i],cOI[i])
        N=_div(pV[i],cV[i])
        O=pOI[i]-cOI[i]
        AN=pP[i]-cP[i]
        P=_div(O,M); Q=_div(L,N); R=_div(O,L); S=_div(L,O); T=_div(P,Q)
        vec=[L,M,N,O,P,Q,R,S,T]
        W=excel_kurt(vec); X=excel_skew(vec)
        Mass=_div(W,X); Force=_div(X,W)
        AP=_div(AN,O); AR=_div(AN,L); AT=_div(AP,AR)
        rows.append(dict(strike=f["strike"].iloc[i],K=K,L=L,M=M,N=N,O=O,AN=AN,
                         P=P,Q=Q,R=R,S=S,T=T,Kurt=W,Skew=X,Mass=Mass,Force=Force,
                         DID=AP,DIT=AR,DR3=AT))
    df=pd.DataFrame(rows)
    # Y = 1 + Kurt_this + Kurt_next ; Z = Y/86400 ; STD_grad = Z_next - Z_this
    kn=df["Kurt"].shift(-1)
    df["ICF"]=1+df["Kurt"]+kn
    df["STD"]=df["ICF"]/86400.0
    df["STD_grad"]=df["STD"].shift(-1)-df["STD"]
    return df

if __name__=="__main__":
    import sys
    sys.path.insert(0,"/mnt/user-data/outputs")
    from quan_engine import ingest_chain
    csv=sys.argv[1] if len(sys.argv)>1 else \
        "/mnt/project/20260413_075705_nqm26optionsmondayweeklyoptionsexp04_13_26showallsidebysideintraday04132026.csv"
    df=per_strike_block(ingest_chain(csv))
    print(df[["strike","Kurt","Skew","Mass","Force","T","P","DID","DR3"]].head(12).to_string(index=False))


# ─────────────────────────────────────────────────────────────
# SCORING LAYER (Playbook Part II) — built on the validated block above.
# ─────────────────────────────────────────────────────────────
OBS = dict(mass=2.0, kurt=4.5, lr=8.0, absA=20.0)        # Layer-1 observable thresholds
# Gradient classification operational params (terminal defaults; flagged for Playbook confirmation)
GRAD = dict(phase_eps=0.5, grad_min=0.5)
DR3_FRESH = 0.3

def observable_scan(df):
    import numpy as np
    f=df.copy()
    f["c_mass"]=f["Mass"].abs()>OBS["mass"]
    f["c_kurt"]=f["Kurt"]>OBS["kurt"]
    f["c_lr"]=f["T"].abs()>OBS["lr"]
    f["c_A"]=f["P"].abs()>OBS["absA"]
    f["criteriaMet"]=f[["c_mass","c_kurt","c_lr","c_A"]].sum(axis=1)
    f["isPDSL"]=f["criteriaMet"]==4
    f["isDSC"]=f["criteriaMet"]==3
    return f

def classify_gradient(df):
    """Playbook Part II Layer 2 — canonical. Force(K) sign + lower-neighbor differential.
       ASCENDING:  F(K)>0 AND F(K)>F(K-1)   (positive and building)
       DESCENDING: F(K)<0 AND F(K)<F(K-1)   (negative and building)
       PHASE BOUNDARY: Force crosses zero between K-1 and K (the 'Force approx 0' reversal zone)
       else: has a sign but not cleanly building -> ambiguous (clear=False)
       No absolute magnitude band (the terminal's 0.5 was incompatible with Force==1/Mass)."""
    import numpy as np
    F=df["Force"].values; n=len(F); types=[]; clears=[]
    for i in range(n):
        fh=F[i] if np.isfinite(F[i]) else 0.0
        fb=F[i-1] if i>0 and np.isfinite(F[i-1]) else fh
        sign_flip = (fh>0) != (fb>0)
        if sign_flip:
            types.append("PHASE_BOUNDARY"); clears.append(True)
        elif fh>0 and fh>fb:
            types.append("ASCENDING"); clears.append(True)
        elif fh<0 and fh<fb:
            types.append("DESCENDING"); clears.append(True)
        else:
            types.append("ASCENDING" if fh>0 else "DESCENDING"); clears.append(False)
    df=df.copy(); df["gradient"]=types; df["gradient_clear"]=clears
    return df

def score_strike(row, realization_dir):
    """5-criterion 0-10. realization_dir in {UP,DOWN,NEUTRAL} feeds the TSC-prior slot."""
    import numpy as np
    pts=0; bd={}
    bd["observable"]=3 if row["criteriaMet"]==4 else 0; pts+=bd["observable"]
    bd["gradient_clear"]=2 if row["gradient_clear"] else 0; pts+=bd["gradient_clear"]
    # TSC prior aligned with gradient — the realization direction (the slot we filled)
    g=row["gradient"]
    if realization_dir=="NEUTRAL" or g=="PHASE_BOUNDARY":
        bd["tsc_prior"]=1
    elif (g=="ASCENDING" and realization_dir=="UP") or (g=="DESCENDING" and realization_dir=="DOWN"):
        bd["tsc_prior"]=2
    else:
        bd["tsc_prior"]=0
    pts+=bd["tsc_prior"]
    dr3=row["DR3"]
    bd["live_dealer"]=2 if (isinstance(dr3,(int,float)) and np.isfinite(dr3) and abs(dr3)<DR3_FRESH) else 0
    pts+=bd["live_dealer"]
    sg=row["STD_grad"]
    bd["icf_rising"]=1 if (isinstance(sg,(int,float)) and np.isfinite(sg) and sg>0) else 0
    pts+=bd["icf_rising"]
    tier=1 if pts>=8 else 2 if pts>=6 else 3 if pts>=4 else 4
    return pts, tier, bd

def scorecard(frame, realization_dir="NEUTRAL", anchor=None):
    df=classify_gradient(observable_scan(per_strike_block(frame)))
    cand=df[df["criteriaMet"]>=3].copy()
    out=[]
    for _,row in cand.iterrows():
        pts,tier,bd=score_strike(row, realization_dir)
        out.append(dict(strike=round(float(row["strike"]),1),
                        kind="PDSL" if row["isPDSL"] else "DSC",
                        score=pts, tier=tier, gradient=row["gradient"],
                        mass=round(float(row["Mass"]),2) if np.isfinite(row["Mass"]) else None,
                        dr3=round(float(row["DR3"]),3) if np.isfinite(row["DR3"]) else None,
                        dist=round(float(row["strike"])-anchor,0) if anchor else None,
                        breakdown=bd))
    out.sort(key=lambda x:(-x["score"], abs(x["dist"]) if x["dist"] is not None else 0))
    return out

import numpy as np


# ─────────────────────────────────────────────────────────────
# CANONICAL TSC PRIOR (Playbook Part I) — the real "+2 TSC prior" source.
# Read from the intent-pressure fold: DIPLTR (Signal 1), ZC quadrant (Signal 2),
# entropy (Signal 3), SOP latent orientation (Signal 4) -> STRONG/MODERATE/NO + dir.
# DIPLTR>0 bullish, <0 bearish (NO inversion — unlike realization's DR3).
# ─────────────────────────────────────────────────────────────
def _num(v): 
    return v if isinstance(v,(int,float)) else float("nan")

def read_tsc_prior(recalced_book):
    import openpyxl, numpy as np
    wb=openpyxl.load_workbook(recalced_book, data_only=True)
    tg=wb["Time State Compass(Gradient)"]; sf=wb["SOP Folding"]
    # Signal 1 — DIPLTR residual at CW +0.9 (row21) and +1.0 (row22); TSC(Gradient) col Q=17
    dip=[_num(tg.cell(r,17).value) for r in (21,22)]; dip=[d for d in dip if np.isfinite(d)]
    dipltr=float(np.mean(dip)) if dip else float("nan")
    dip_dir="UP" if dipltr>0 else "DOWN" if dipltr<0 else "NEUTRAL"
    # Signal 4 — SOP latent (SOPG=Q17, SOPC path=R18) at CW +0.7/+0.8/+0.9 (rows 19,20,21)
    qv=[_num(sf.cell(r,17).value) for r in (19,20,21)]; rv=[_num(sf.cell(r,18).value) for r in (19,20,21)]
    qm=np.nanmean(qv); rm=np.nanmean(rv)
    sop_dir=("UP" if (qm>0 and rm>0) else "DOWN" if (qm<0 and rm<0) else "CONFLICTED")
    # Signal 2 — zero-cross count + final ZC quadrant (SOP Folding S=19 flag, A=1 CW)
    zc_rows=[r for r in range(2,23) if _num(sf.cell(r,19).value)==1]
    zc_count=len(zc_rows)
    final_cw=_num(sf.cell(zc_rows[-1],1).value) if zc_rows else float("nan")
    zc_confirm = (np.isfinite(final_cw) and final_cw>=0.5 and dip_dir!="NEUTRAL")
    # Signal 3 — entropy (sum V=22); ZC count as turbulence proxy
    ent=[_num(sf.cell(r,22).value) for r in range(2,23)]
    entropy_total=float(np.nansum(ent))
    high_entropy = zc_count>=5
    # Synthesis (Part I step 05): directional votes from DIPLTR, SOP latent, ZC-confirm
    votes={"UP":0,"DOWN":0}
    if dip_dir in votes: votes[dip_dir]+=1
    if sop_dir in votes: votes[sop_dir]+=1
    if zc_confirm and dip_dir in votes: votes[dip_dir]+=1
    dom=max(votes,key=votes.get); agree=votes[dom]
    if votes["UP"]==votes["DOWN"]:
        prior_dir, strength="NEUTRAL","NO"
    else:
        prior_dir=dom
        strength = "STRONG" if agree>=3 else "MODERATE" if agree>=2 else "NO"
        if strength=="NO": prior_dir="NEUTRAL"
    if high_entropy and strength=="STRONG": strength="MODERATE"   # entropy humility downgrade
    # Position-arc completeness: Part I reads CW +0.7..+1.0 (rows 19-22). A single-snapshot
    # fold only populates the negation arc (CW -1..0, rows 2-12); the position arc fills LIVE.
    pos_cells=[ _num(sf.cell(r,c).value) for r in (19,20,21,22) for c in (10,15,17,18) ]
    arc_complete = any(np.isfinite(v) and v!=0 for v in pos_cells)
    if not arc_complete:
        prior_dir, strength = "NEUTRAL", "NO_ARC"   # canonical prior needs the live close arc
    return dict(prior_dir=prior_dir, strength=strength, dipltr=round(dipltr,4), dip_dir=dip_dir,
                sop_dir=sop_dir, zc_count=zc_count, final_zc_cw=final_cw,
                entropy=round(entropy_total,4), high_entropy=high_entropy,
                arc_complete=arc_complete)
