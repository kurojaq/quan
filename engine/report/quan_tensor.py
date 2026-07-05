"""
quan_tensor.py — canonical tensor surface per session, via the parser path.

Replicates parser.py's inject_csv_into_template (CSV -> golden_reference Book by
header name, CSV row order), recalcs the full workbook with LibreOffice headless
(which propagates chain -> cascade -> pressure -> TSC -> SOP fold -> Tensor ->
Velocity -> Jerk), and reads the canonical surface. No hand-port of the fold; the
spreadsheet's own formulas compute it. Validated against nq_0413_Book.
"""
import os, sys, shutil, subprocess, glob, json
import pandas as pd
import openpyxl

TEMPLATE = "/home/claude/golden_reference.xlsx"

CSV_TO_BOOK_RENAME = {  # verbatim from parser.py
    "strike": "strike",
    "premium": "call premium", "premium.1": "put premium",
    "openint": "call oi",      "openint.1": "put oi",
    "volume": "call vol",      "volume.1": "put vol",
    "latest": "call latest",   "latest.1": "put latest",
}

def _norm(c):
    return str(c).strip().lower().replace(" ", "").replace("_", "")

def _strip_sec(v):
    return v[:-1] if isinstance(v, str) and v.lower().endswith("s") else v


def _cc_cd_atm(frame, anchor, source="intent"):
    """Port of terminal sampleCWPressureValues: 21-pt pressure gradient/curvature
    centered on ATM. BH=|AP| (the (DB37-DB38)*DB31 multiplier cancels exactly in
    the BI [0,1] normalization, so dropping it is identity for all downstream)."""
    import numpy as np
    f = frame.copy()
    f["strike"] = pd.to_numeric(f["strike"], errors="coerce")
    f = f.dropna(subset=["strike"]).sort_values("strike").reset_index(drop=True)
    s = f["strike"].values
    atm = int(np.abs(s - anchor).argmin())
    netOI   = f["putOI"].fillna(0).values   - f["callOI"].fillna(0).values
    netPrem = f["putPrem"].fillna(0).values  - f["callPrem"].fillna(0).values
    netTxn  = f["putVol"].fillna(0).values   - f["callVol"].fillna(0).values
    if source == "realization":            # AT = DR3 = NetTxn / NetOI (the realized flow)
        X = np.where((netOI != 0) & (netTxn != 0), netTxn / np.where(netOI==0,1,netOI), np.nan)
    elif source == "transaction":          # AR = DIT = NetPrem / NetTxn
        X = np.where((netTxn != 0) & (netPrem != 0), netPrem / np.where(netTxn==0,1,netTxn), np.nan)
    else:                                   # intent  AP = DID = NetPrem / NetOI
        X = np.where((netOI != 0) & (netPrem != 0), netPrem / np.where(netOI==0,1,netOI), np.nan)
    BH = np.abs(X)
    valid = BH[np.isfinite(BH)]
    if valid.size < 5: return None
    lo, hi = valid.min(), valid.max(); rng = hi - lo
    BI = np.where(np.isfinite(BH), (BH - lo)/rng if rng>0 else 0.0, np.nan)
    N = len(BI)
    BN = np.full(N-1, np.nan)
    for i in range(N-1):
        if np.isfinite(BI[i]) and np.isfinite(BI[i+1]): BN[i] = BI[i+1]-BI[i]
    start = atm - 10
    cc = np.zeros(21)
    for i in range(21):
        a, b = start+i, start+i+1
        if 0 <= a and b < len(BN) and np.isfinite(BN[a]) and np.isfinite(BN[b]):
            cc[i] = (BN[b]-BN[a]) / 0.1
    cd = np.zeros(21)
    for i in range(20): cd[i] = (cc[i+1]-cc[i]) / 0.1
    return cc, cd, dict(atm_strike=float(s[atm]), atm_idx=atm, n=N,
                        window_covered=bool(start>=0 and start+21 < len(BN)))


def inject_pressure(frame, anchor, out_path, source="intent"):
    """Write ATM-centered cc/cd into Book CC(81)/CD(82) rows 2-22; canonical
    TSC->fold->tensor recalcs from there. Leaves all formulas else intact."""
    res = _cc_cd_atm(frame, anchor, source)
    if res is None: return None
    cc, cd, meta = res
    shutil.copy(TEMPLATE, out_path)
    wb = openpyxl.load_workbook(out_path, data_only=False)
    bk = wb["Book"]
    for i in range(21):
        bk.cell(2+i, 81).value = float(cc[i])   # CC = Pressure Gradient
        bk.cell(2+i, 82).value = float(cd[i])   # CD = Pressure Curvature
    wb.save(out_path); return meta


def recalc(path, outdir=None):
    # Use the xlsx skill's recalc (forces a true formula recompute, edits in place).
    # The raw `soffice --convert-to xlsx` only reformats — it does NOT recalculate,
    # which left Tensor/Jerk/SOP sheets stale/empty. This computes them.
    import subprocess
    r = subprocess.run(["python3", "/mnt/skills/public/xlsx/scripts/recalc.py", path, "60"],
                       capture_output=True, text=True, timeout=240)
    if r.returncode != 0:
        raise RuntimeError("recalc failed: " + (r.stderr or r.stdout)[-300:])
    return path   # recalc.py edits the file in place

def extract(recalced_xlsx):
    wb = openpyxl.load_workbook(recalced_xlsx, data_only=True)
    ts, sj, sf = wb["Tensor Surface"], wb["Surface Jerk"], wb["SOP Folding"]
    offsets = [ts.cell(1, c).value for c in range(2, 203)]
    def prof(ws, r):
        v = [ws.cell(r, c).value for c in range(2, 203)]
        return [x if isinstance(x, (int, float)) else 0.0 for x in v]
    rows = []
    for r in range(2, 14):
        p = prof(ts, r); s = sum(p)
        pk = offsets[max(range(len(p)), key=lambda k: p[k])] if s > 0 else None
        left = sum(p[k] for k in range(len(p)) if offsets[k] < 0)
        right = sum(p[k] for k in range(len(p)) if offsets[k] > 0)
        asym = (right - left) / (right + left) if (right + left) > 0 else 0.0
        O, Q, A = sf.cell(r, 15).value, sf.cell(r, 17).value, sf.cell(r, 1).value
        jp = prof(sj, r); jerk = sum(abs(x) for x in jp)
        joff = offsets[max(range(len(jp)), key=lambda k: abs(jp[k]))] if any(jp) else None
        rows.append(dict(chronoT=A, ampO=abs(O) if isinstance(O,(int,float)) else 0.0,
                         Q=Q, rowSum=s, asym=asym, peak=pk, jerk=jerk, jerk_off=joff))
    active = [r for r in rows if r["rowSum"] > 1e-9]
    peaks = [r["peak"] for r in active if r["peak"] is not None]
    asyms = [r["asym"] for r in active]
    migr = (max(peaks) - min(peaks)) if peaks else 0.0
    mean_asym = sum(asyms)/len(asyms) if asyms else 0.0
    geometry = "COMPRESSION" if (peaks and abs(sum(peaks)/len(peaks)) < 5 and migr < 10) else "TRAJECTORY"
    direction = "UP" if mean_asym > 0.05 else "DOWN" if mean_asym < -0.05 else "NEUTRAL"
    # peak-jerk row = where instability concentrates (release timing)
    jrow_r = max(active, key=lambda r: r["jerk"]) if active else None
    arow_r = max(active, key=lambda r: r["ampO"]) if active else None
    jrow = jrow_r["chronoT"] if jrow_r else None
    arow = arow_r["chronoT"] if arow_r else None
    amp_off = arow_r["peak"] if arow_r else None
    jerk_off = jrow_r["jerk_off"] if jrow_r else None
    return dict(rows=rows, geometry=geometry, direction=direction,
                mean_asym=round(mean_asym,4), peak_migration=round(migr,3),
                amp_peak_chronoT=arow, jerk_peak_chronoT=jrow,
                amp_peak_offset=amp_off, jerk_peak_offset=jerk_off,
                active_rows=len(active))


import numpy as _np
def _atm_center(df, anchor):
    """Reorder ascending chain so the ATM strike lands at Book data-row 25
    (so the hard-coded CC anchors rows15-35 == ATM-10..ATM+10). We keep the
    full ascending chain but choose the row offset; if not enough strikes below
    ATM, we pad by shifting (the canonical pad-with-zero behavior handles edges)."""
    strikes = pd.to_numeric(df["strike"].astype(str).str.replace(",","",regex=False), errors="coerce")
    atm_pos = int((strikes - anchor).abs().values.argmin())
    # We want atm_pos to map to data-row 25 => data-row index 23 (0-based, since row2=index0).
    target = 23
    shift = target - atm_pos
    if shift > 0:
        pad = pd.DataFrame([{c: None for c in df.columns}] * shift)
        df2 = pd.concat([pad, df], ignore_index=True)
    elif shift < 0:
        df2 = df.iloc[-shift:].reset_index(drop=True)
    else:
        df2 = df.copy()
    return df2

def inject_centered(csv_path, out_path, anchor):
    shutil.copy(TEMPLATE, out_path)
    wb = openpyxl.load_workbook(out_path, data_only=False)
    ws = wb["Book"]
    df = pd.read_csv(csv_path)
    df.columns = [_norm(c) for c in df.columns]
    rmap = {_norm(k): _norm(v) for k, v in CSV_TO_BOOK_RENAME.items()}
    df = df.rename(columns=rmap)
    df = df[list(rmap.values())]
    for col in ["calllatest", "putlatest"]:
        if col in df.columns:
            df[col] = df[col].apply(_strip_sec).apply(pd.to_numeric, errors="coerce")
    df = _atm_center(df, anchor)
    hdr = next(r for r in range(1, 15) if any(ws.cell(r, c).value for c in range(1, 30)))
    bidx = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(hdr, c).value
        if v is not None:
            bidx[_norm(v)] = c
    for ri, row in enumerate(df.itertuples(index=False), start=hdr + 1):
        for name, val in zip(df.columns, row):
            if val is not None and (isinstance(val,str) or not (isinstance(val,float) and _np.isnan(val))):
                ws.cell(ri, bidx[name]).value = val
    wb.save(out_path)
    return len(df)

SESSIONS = {
  "Apr-13": "/mnt/project/20260413_075705_nqm26optionsmondayweeklyoptionsexp04_13_26showallsidebysideintraday04132026.csv",
  "May-14": "/mnt/project/nqm26optionsfridayweeklyoptionsexp05_15_26showallsidebysideintraday05142026.csv",
  "May-15": "/mnt/project/nqm26optionsfridayweeklyoptionsexp05_15_26showallsidebysideintraday05152026.csv",
  "May-18": "/mnt/project/nqm26optionstuesdayweeklyoptionsexp05_19_26showallsidebysideintraday05182026.csv",
  "May-21": "/mnt/project/nqm26optionsthursdayweeklyoptionsexp05_21_26showallsidebysideintraday05212026_1.csv",
}

ANCHORS = {"Apr-13":25216,"May-14":29476,"May-15":29278,"May-18":29338,"May-21":29217}

ANCHORS = {"Apr-13":25216,"May-14":29476,"May-15":29278,"May-18":29338,"May-21":29217}

def run_one(label, csv, source="intent"):
    import sys; sys.path.insert(0,'/mnt/user-data/outputs')
    from quan_engine import ingest_chain
    book = f"/home/claude/_t_{label}.xlsx"
    meta = inject_pressure(ingest_chain(csv), ANCHORS[label], book, source)
    if meta is None: return dict(label=label, geometry="EMPTY", direction="NA",
                                 mean_asym=0, peak_migration=0, amp_peak_chronoT=None,
                                 jerk_peak_chronoT=None, active_rows=0, n_strikes=0)
    rec = recalc(book, f"/home/claude/_rc_{label}_{source}")
    st = extract(rec); st["label"]=label; st["n_strikes"]=meta["n"]; st["atm_covered"]=meta["window_covered"]
    return st

if __name__ == "__main__":
    only = sys.argv[1] if len(sys.argv) > 1 else None
    for label, csv in SESSIONS.items():
        if only and label != only: continue
        st = run_one(label, csv)
        print(f"\n=== {label}  (window covered={st.get('atm_covered')}, chain n={st.get('n_strikes')}) ===")
        print(f"  geometry={st['geometry']}  direction={st['direction']} "
              f"(asym {st['mean_asym']:+.3f})  migration={st['peak_migration']}")
        print(f"  amp peak @ chronoT {st['amp_peak_chronoT']}  jerk peak @ {st['jerk_peak_chronoT']}  active_rows={st['active_rows']}")


def session_tensor(csv_path, anchor, source="intent"):
    """Session-level tensor read for ANY anchor (not just the hardcoded ANCHORS dict).
    Returns the extract dict PLUS tphase: the release timing as a 0-1 session fraction
    (chronoT runs -1=open .. +1=close, so tphase = (chronoT+1)/2) for the execution TPHASE gate."""
    import sys, tempfile, os
    sys.path.insert(0, "/mnt/user-data/outputs")
    from quan_engine import ingest_chain
    book = os.path.join(tempfile.gettempdir(), "tensor_session.xlsx")
    meta = inject_pressure(ingest_chain(csv_path), anchor, book, source)
    if meta is None:
        return None
    recalc(book)
    st = extract(book)
    jt = st.get("jerk_peak_chronoT")
    st["tphase"] = round((jt + 1) / 2, 3) if isinstance(jt, (int, float)) else None
    st["atm_covered"] = meta.get("window_covered")
    return st
